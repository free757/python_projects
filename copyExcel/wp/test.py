import time
import sys

import openpyxl
import  os

import pandas

'''l=[]
rows=56
columns=3
for i in range(1,rows+1):
    l.append([])
    print(i)

print(l)





toolbar_width = 40

# setup toolbar
sys.stdout.write("[%s]" % (" " * toolbar_width))
sys.stdout.flush()
sys.stdout.write("\b" * (toolbar_width+1)) # return to start of line, after '['

for i in range(toolbar_width):
    time.sleep(0.1) # do real work here
    # update the bar
    sys.stdout.write("-")
    sys.stdout.flush()

sys.stdout.write("]\n") # this ends the progress bar
'''
'''wb_1=openpyxl.load_workbook('C:/Users/free\PycharmProjects\copyExcel/wp/files/FW_and_IR_Throughput_Link_Usage_-_3G-Gi-Total-MX480-LDN_Overtime.xlsx')
sheet_1_wb_1=wb_1.get_sheet_by_name('Sheet1')
maxR=sheet_1_wb_1.max_row
print(maxR)
'''



'''def getListOfFiles(dirName):
    # create a list of file and sub directories
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)

    return allFiles
allfiles=list()
#(i) is (folders in p "subDir" )
for i in p:
    fpath='C:/Users/free/PycharmProjects/copyExcel/wp/files/FW_and_IR_Throughput_2019.08.25 at 08_0124 EET/'+i+'/4444.xlsx'
    allfiles.append(fpath)
print(allfiles)
for f in allfiles:
    wb1=openpyxl.load_workbook(f)
    data=wb1['Sheet1']['A1'].value
    print(data)


fullPath = 'C:/Users/free/PycharmProjects/copyExcel/wp/files/FW_and_IR_Throughput_2019.08.25 at 08_0124 EET/New folder'

print(fullPath)
g=os.listdir(fullPath)
print(g)'''
#converting , appending to lists
root_path='E:/Work from home/shared/'
firewall_csv_file_name='/FW_and_IR_Throughput_Link_Usage_-_DEV_Total_Gi_Traffic_Overtime.csv'
router_csv_file_name='/FW_and_IR_Throughput_Link_Usage_-_3G-Gi-Total-MX480-LDN_Overtime.csv'
firewall_xlsx_file_name='/FW_and_IR_Throughput_Link_Usage_-_DEV_Total_Gi_Traffic_Overtime.xlsx'
router_xlsx_file_name='/FW_and_IR_Throughput_Link_Usage_-_3G-Gi-Total-MX480-LDN_Overtime.xlsx'
p=os.listdir(root_path)
print(p)

#lists to append>>

router_filesList=list()
firewall_filesList=list()
for i in p:
    #router files path
    router_csv_filesPath=root_path+i+router_csv_file_name
    router_xlsx_filesPath=root_path+i+router_xlsx_file_name
    #firewall files path
    firewall_csv_filesPath=root_path+i+firewall_csv_file_name
    firewall_xlsx_filesPath=root_path+i+firewall_xlsx_file_name
    #converting router files>>> csv to xlsx
    read_router_csvFiles=pandas.read_csv(router_csv_filesPath)
    read_router_csvFiles.to_excel(router_xlsx_filesPath)
    #converting firewall files>>> csv to xlsx
    read_firewall_csvFiles=pandas.read_csv(firewall_csv_filesPath)
    read_firewall_csvFiles.to_excel(firewall_xlsx_filesPath)
    #appinding the new files to router_filesList
    router_filesList.append(router_xlsx_filesPath)
    #appinding the new files to firewall_filesList

    firewall_filesList.append(firewall_xlsx_filesPath)

print(router_filesList)
print(firewall_filesList)


