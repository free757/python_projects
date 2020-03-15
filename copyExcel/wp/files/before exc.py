import pandas
import os
import openpyxl
import xlsx
import numpy
####

print('hello')

#wb>> which copy inside it
print('loading>.........wb_2')
wb_2=openpyxl.load_workbook('E:/Work from home/Done/Firewall vs (Autosaved).xlsx')
print('done')

####fristly we need to convert csv fil to xlsx file

#converting , appending to lists

print('converting>>>>')
root_path='E:/Work from home/shared/'
firewall_csv_file_name='/FW_and_IR_Throughput_Link_Usage_-_DEV_Total_Gi_Traffic_Overtime.csv'
router_csv_file_name='/FW_and_IR_Throughput_Link_Usage_-_3G-Gi-Total-MX480-LDN_Overtime.csv'
firewall_xlsx_file_name='/FW_and_IR_Throughput_Link_Usage_-_DEV_Total_Gi_Traffic_Overtime.xlsx'
router_xlsx_file_name='/FW_and_IR_Throughput_Link_Usage_-_3G-Gi-Total-MX480-LDN_Overtime.xlsx'
p=os.listdir(root_path)
print(p)

#lists to append>>

router_filesList=list()
firewall_filesList=list()
for i in p:
    #router files path
    router_csv_filesPath=root_path+i+router_csv_file_name
    router_xlsx_filesPath=root_path+i+router_xlsx_file_name
    #firewall files path
    firewall_csv_filesPath=root_path+i+firewall_csv_file_name
    firewall_xlsx_filesPath=root_path+i+firewall_xlsx_file_name
    #converting router files>>> csv to xlsx
    read_router_csvFiles=pandas.read_csv(router_csv_filesPath)
    read_router_csvFiles.to_excel(router_xlsx_filesPath)
    #converting firewall files>>> csv to xlsx
    read_firewall_csvFiles=pandas.read_csv(firewall_csv_filesPath)
    read_firewall_csvFiles.to_excel(firewall_xlsx_filesPath)
    print('coverting DONE')
    #appinding the new files to router_filesList
    router_filesList.append(router_xlsx_filesPath)
    #appinding the new files to firewall_filesList

    firewall_filesList.append(firewall_xlsx_filesPath)

print(router_filesList)
print(firewall_filesList)



###load workbooks which are copy from to
for rf in router_filesList:
    print('loading wb_1>/......')
    wb_1=openpyxl.load_workbook(rf)
    print('done')

    #select which sheets inside workbook
    sheet_1_wb_1=wb_1.get_sheet_by_name('Sheet1')
    sheet_1_wb_2=wb_2.get_sheet_by_name('Router')
    ##catch data to copy it
    l=[]
    rows_sh_wb1=sheet_1_wb_1.max_row
    columns=3
    for i in range(1,rows_sh_wb1):
        l.append([])

    print(l)

    ##
    for r in range(2,rows_sh_wb1+1):
        for c in range(1,columns+1):
            cdata=sheet_1_wb_1.cell(row=r,column=c)
            l[r-2].append(cdata.value)
    print(l)
    maxR_sh_wb2=sheet_1_wb_2.max_row
    print(maxR_sh_wb2)
    imaxR_sh_wb2=maxR_sh_wb2+rows_sh_wb1 -1
    print(imaxR_sh_wb2)

    for r in range(maxR_sh_wb2,imaxR_sh_wb2):
        for c in range(1,columns+1):
            putData=sheet_1_wb_2.cell(row=r+1,column=c)
            putData.value=l[r-maxR_sh_wb2][c-1]
            print('___')

print("saving......")

wb_2.save('E:/Work from home/Done/Firewall vs (Autosaved).xlsx')

print("done")
